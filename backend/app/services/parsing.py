"""Läsning och tolkning av bankexporter (CSV/XLSX) med svenska egenheter.

Pipeline:
  inspect_file(bytes, filename)  -> InspectionResult (encoding, delimiter,
                                    headerrad, fingerprint, gissad mappning)
  parse_with_options(bytes, ...) -> ParseResult med rader i ören/ISO-datum
"""
from __future__ import annotations

import csv
import io
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from .fingerprint import compute_fingerprint
from .normalize import normalize_description

DATE_FORMATS = ["%Y-%m-%d", "%Y/%m/%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y%m%d"]
ENCODINGS = ["utf-8-sig", "utf-8", "cp1252"]
DELIMITERS = [";", ",", "\t"]

_DATE_RE = re.compile(r"^\s*\d{1,4}[-./]\d{1,2}[-./]\d{1,4}\s*$|^\s*\d{8}\s*$")
_AMOUNT_RE = re.compile(r"^\s*[+-−]?[\d\s\xa0.,]+\s*(kr)?\s*$")


@dataclass
class ParseOptions:
    file_type: str                      # 'csv' | 'xlsx'
    column_mapping: dict                # {"date": i, "description": i, "amount": i|None,
                                        #  "amount_in": i|None, "amount_out": i|None,
                                        #  "balance": i|None, "merchant": i|None}
    delimiter: str | None = ";"
    encoding: str | None = "utf-8-sig"
    decimal_separator: str = ","
    thousands_separator: str | None = None
    date_format: str = "%Y-%m-%d"
    header_row_index: int = 0
    invert_sign: bool = False
    skip_value: str | None = None

    @classmethod
    def from_profile(cls, p) -> "ParseOptions":
        return cls(
            file_type=p.file_type,
            column_mapping=json.loads(p.column_mapping),
            delimiter=p.delimiter,
            encoding=p.encoding,
            decimal_separator=p.decimal_separator,
            thousands_separator=p.thousands_separator,
            date_format=p.date_format,
            header_row_index=p.header_row_index,
            invert_sign=bool(p.invert_sign),
            skip_value=p.skip_value,
        )


@dataclass
class ParsedRow:
    booked_date: str
    amount_ore: int
    description_raw: str
    description_norm: str
    balance_ore: int | None
    row_index: int
    member: str | None = None


@dataclass
class ParseResult:
    rows: list[ParsedRow] = field(default_factory=list)
    skipped: list[dict] = field(default_factory=list)   # {"row_index", "reason", "cells"}


@dataclass
class InspectionResult:
    file_type: str
    encoding: str | None
    delimiter: str | None
    header_row_index: int
    header: list[str]
    sample_rows: list[list[str]]
    fingerprint: str
    suggested_mapping: dict
    suggested_date_format: str
    suggested_decimal_separator: str
    suggested_thousands_separator: str | None
    suggested_invert_sign: bool


# ---------------------------------------------------------------- inläsning

def read_raw_rows(data: bytes, filename: str) -> tuple[str, list[list], str | None, str | None]:
    """-> (file_type, rader med råceller, encoding, delimiter)

    För PDF används 'delimiter'-fältet till utfärdaridentiteten (entercard,
    amex …) så att olika korts fakturor får olika formatprofiler.
    """
    if filename.lower().endswith(".pdf") or data[:5] == b"%PDF-":
        rows, issuer = _read_pdf(data)
        return ("pdf", rows, None, issuer)
    if filename.lower().endswith((".xlsx", ".xlsm", ".xls")):
        return ("xlsx", _read_xlsx(data), None, None)
    encoding = _detect_encoding(data)
    text = data.decode(encoding)
    delimiter = _detect_delimiter(text)
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    return ("csv", [row for row in reader], encoding, delimiter)


_PDF_DATE = r"\d{4}-\d{2}-\d{2}|\d{2}[-./]\d{2}[-./]\d{2,4}"
_PDF_ROW = re.compile(
    rf"^({_PDF_DATE})(?:\s+({_PDF_DATE}))?\s+(.+?)\s+(-?\d[\d\s\xa0.]*,\d{{2}}-?)$"
)
_PDF_ISSUERS = (
    ("entercard", "entercard"),
    ("re:member", "entercard"),
    ("american express", "amex"),
    ("swedbank", "swedbank"),
    ("handelsbanken", "handelsbanken"),
)


def _read_pdf(data: bytes) -> tuple[list[list], str]:
    """Extrahera transaktionsrader ur en kortfaktura-PDF.

    Heuristik: rader som börjar med en–två datum och slutar med ett belopp
    med decimalkomma. Fungerar för Entercard- och Amex-fakturor m.fl.
    """
    import pdfplumber

    pages = []
    with pdfplumber.open(io.BytesIO(data)) as pdf:
        for page in pdf.pages:
            pages.append(page.extract_text() or "")
    all_text = "\n".join(pages)

    issuer = "pdf"
    low = all_text.casefold()
    for needle, name in _PDF_ISSUERS:
        if needle in low:
            issuer = name
            break

    rows: list[list] = [["Transaktionsdatum", "Bokföringsdatum", "Beskrivning", "Belopp"]]
    for line in all_text.splitlines():
        m = _PDF_ROW.match(line.strip())
        if not m:
            continue
        d1, d2, desc, amount = m.groups()
        rows.append([d1, d2 or d1, desc.strip(), amount])
    if len(rows) == 1:
        raise ValueError(
            "Hittade inga transaktionsrader i PDF:en — är det en kortfaktura med tabell?"
        )
    return rows, issuer


def _read_xlsx(data: bytes) -> list[list]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = [[c for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _detect_encoding(data: bytes) -> str:
    if data.startswith(b"\xef\xbb\xbf"):
        return "utf-8-sig"
    if data.startswith(b"\xff\xfe") or data.startswith(b"\xfe\xff"):
        return "utf-16"  # annars blir UTF-16 NUL-interfolierat cp1252-skräp
    for enc in ("utf-8", "cp1252"):
        try:
            data.decode(enc)
            return enc
        except UnicodeDecodeError:
            continue
    return "latin-1"  # kan aldrig misslyckas — hellre läsbart än HTTP-fel


def _detect_delimiter(text: str) -> str:
    lines = [ln for ln in text.splitlines()[:20] if ln.strip()]
    best, best_score = ";", -1.0
    for d in DELIMITERS:
        counts = [ln.count(d) for ln in lines]
        positive = [c for c in counts if c > 0]
        if not positive:
            continue
        # konsistens: många rader med samma antal avgränsare
        from statistics import median
        med = median(positive)
        consistent = sum(1 for c in positive if c == med)
        score = consistent * med
        if score > best_score:
            best, best_score = d, score
    return best


# ------------------------------------------------------------- celltolkning

def _cell_str(c) -> str:
    if c is None:
        return ""
    if isinstance(c, (datetime, date)):
        return c.strftime("%Y-%m-%d")
    return str(c).strip()


def looks_like_date(c) -> bool:
    if isinstance(c, (datetime, date)):
        return True
    return bool(_DATE_RE.match(_cell_str(c)))


def looks_like_amount(c) -> bool:
    if isinstance(c, (int, float)) and not isinstance(c, bool):
        return True
    s = _cell_str(c)
    return bool(s) and bool(_AMOUNT_RE.match(s)) and any(ch.isdigit() for ch in s)


def parse_amount(cell, decimal_sep: str = ",", thousands_sep: str | None = None) -> int:
    """Belopp → ören (int). Kastar ValueError vid obegripligt innehåll."""
    if isinstance(cell, (int, float)) and not isinstance(cell, bool):
        return round(Decimal(str(cell)) * 100).__int__()
    s = _cell_str(cell).replace("\xa0", " ").replace("−", "-").replace("kr", "").strip()
    if not s:
        raise ValueError("tomt belopp")
    # negativa format som vissa banker använder: "(1 234,00)" och "1 234,00-"
    negative = False
    if s.startswith("(") and s.endswith(")"):
        negative = True
        s = s[1:-1].strip()
    if s.endswith("-"):
        negative = True
        s = s[:-1].strip()
    if thousands_sep:
        s = s.replace(thousands_sep, "")
    s = s.replace(" ", "")
    if decimal_sep == ",":
        s = s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", "")
    try:
        value = int(round(Decimal(s) * 100))
    except InvalidOperation as e:
        raise ValueError(f"kan inte tolka belopp: {cell!r}") from e
    return -abs(value) if negative else value


def parse_date_cell(cell, fmt: str) -> str:
    if isinstance(cell, datetime):
        return cell.date().isoformat()
    if isinstance(cell, date):
        return cell.isoformat()
    s = _cell_str(cell)
    return datetime.strptime(s, fmt).date().isoformat()


def detect_date_format(samples: list) -> str:
    strs = [_cell_str(s) for s in samples if s is not None and not isinstance(s, (datetime, date))]
    strs = [s for s in strs if s]
    if not strs:
        return "%Y-%m-%d"
    for fmt in DATE_FORMATS:
        ok = 0
        for s in strs:
            try:
                datetime.strptime(s, fmt)
                ok += 1
            except ValueError:
                pass
        if ok and ok >= len(strs) * 0.6:
            return fmt
    return "%Y-%m-%d"


# ----------------------------------------------------------- headerdetektering

def detect_header_row(rows: list[list]) -> int:
    """Första raden med ≥3 textceller där nästa icke-tomma rad ser ut som data."""
    for i, row in enumerate(rows[:15]):
        cells = [_cell_str(c) for c in row]
        texty = [
            c for c in cells
            if c and not looks_like_date(c) and not looks_like_amount(c)
        ]
        if len(texty) < 3:
            continue
        for nxt in rows[i + 1 : i + 4]:
            nxt_cells = [c for c in nxt if _cell_str(c)]
            if not nxt_cells:
                continue
            if any(looks_like_date(c) for c in nxt) and any(looks_like_amount(c) for c in nxt):
                return i
            break
    return 0


# ------------------------------------------------------------- mappningsgissning

_HEADER_KEYWORDS: list[tuple[str, tuple[str, ...]]] = [
    ("date", ("bokföringsdag", "bokföringsdatum", "bokförd", "transaktionsdatum",
              "transaktionsdag", "datum", "date", "reskontradatum")),
    ("description", ("beskrivning", "text/mottagare", "text", "specifikation", "namn",
                     "rubrik", "meddelande", "description", "värdepapper/beskrivning")),
    ("amount", ("belopp", "amount", "summa")),
    ("amount_in", ("insättning", "insättningar", "inbetalning", "kredit")),
    ("amount_out", ("uttag", "utbetalning", "debet")),
    ("balance", ("bokfört saldo", "saldo", "balance")),
    ("member", ("kortmedlem", "kortinnehavare", "ägare", "innehavare", "medlem", "cardmember")),
]


def guess_mapping(header: list[str], data_rows: list[list]) -> dict:
    mapping: dict = {"date": None, "description": None, "amount": None,
                     "amount_in": None, "amount_out": None, "balance": None, "member": None}
    lower = [_cell_str(h).casefold() for h in header]
    for fld, keywords in _HEADER_KEYWORDS:
        for kw in keywords:
            for i, h in enumerate(lower):
                if h == kw and mapping[fld] is None and i not in _claimed(mapping):
                    mapping[fld] = i
                    break
            if mapping[fld] is not None:
                break
    # kolumninnehålls-sniffning som fallback
    if mapping["date"] is None:
        for i in range(len(header)):
            col = [r[i] for r in data_rows if i < len(r)]
            if col and sum(looks_like_date(c) for c in col) >= len(col) * 0.6:
                mapping["date"] = i
                break
    if mapping["amount"] is None and mapping["amount_in"] is None:
        for i in range(len(header)):
            if i in _claimed(mapping):
                continue
            col = [r[i] for r in data_rows if i < len(r) and _cell_str(r[i])]
            if col and all(looks_like_amount(c) for c in col):
                mapping["amount"] = i
                break
    if mapping["description"] is None:
        for i in range(len(header)):
            if i in _claimed(mapping):
                continue
            col = [_cell_str(r[i]) for r in data_rows if i < len(r)]
            if col and sum(1 for c in col if c and not looks_like_date(c) and not looks_like_amount(c)) >= len(col) * 0.5:
                mapping["description"] = i
                break
    return mapping


def _claimed(mapping: dict) -> set[int]:
    return {v for v in mapping.values() if v is not None}


# ------------------------------------------------------------------ inspect

def inspect_file(data: bytes, filename: str) -> InspectionResult:
    file_type, rows, encoding, delimiter = read_raw_rows(data, filename)
    header_idx = detect_header_row(rows)
    header = [_cell_str(c) for c in rows[header_idx]]
    data_rows = [r for r in rows[header_idx + 1 :] if any(_cell_str(c) for c in r)]
    samples = data_rows[:10]
    mapping = guess_mapping(header, samples)

    date_samples = [r[mapping["date"]] for r in samples
                    if mapping["date"] is not None and mapping["date"] < len(r)]
    amount_col = mapping["amount"] if mapping["amount"] is not None else mapping["amount_out"]
    amount_samples = [_cell_str(r[amount_col]) for r in samples
                      if amount_col is not None and amount_col < len(r)]
    decimal_sep = _guess_decimal_separator(amount_samples)
    thousands = _guess_thousands_separator(amount_samples, decimal_sep)

    # kreditkortsheuristik: om nästan alla belopp är positiva och det ser ut som
    # köp (ingen saldokolumn) är det troligen ett kort som listar köp positivt
    invert = False
    if mapping["amount"] is not None and mapping["balance"] is None and amount_samples:
        parsed = []
        for s in amount_samples:
            try:
                parsed.append(parse_amount(s, decimal_sep, thousands))
            except ValueError:
                pass
        if len(parsed) >= 2 and sum(1 for v in parsed if v > 0) >= len(parsed) * 0.6:
            invert = True

    return InspectionResult(
        file_type=file_type,
        encoding=encoding,
        delimiter=delimiter,
        header_row_index=header_idx,
        header=header,
        sample_rows=[[_cell_str(c) for c in r] for r in samples],
        fingerprint=compute_fingerprint(file_type, delimiter, header),
        suggested_mapping=mapping,
        suggested_date_format=detect_date_format(date_samples),
        suggested_decimal_separator=decimal_sep,
        suggested_thousands_separator=thousands,
        suggested_invert_sign=invert,
    )


def _guess_decimal_separator(samples: list[str]) -> str:
    comma = sum(1 for s in samples if re.search(r",\d{1,2}$", s.strip()))
    point = sum(1 for s in samples if re.search(r"\.\d{1,2}$", s.strip()))
    return "." if point > comma else ","


def _guess_thousands_separator(samples: list[str], decimal_sep: str) -> str | None:
    joined = " ".join(samples).replace("\xa0", " ")
    if re.search(r"\d [\d]{3}", joined):
        return " "
    other = "." if decimal_sep == "," else ","
    if re.search(rf"\d\{other}\d{{3}}", joined):
        return other
    return None


# -------------------------------------------------------------------- parse

def parse_with_options(data: bytes, filename: str, opts: ParseOptions) -> ParseResult:
    _, rows, _, _ = read_raw_rows(data, filename)
    result = ParseResult()
    m = opts.column_mapping

    for idx, row in enumerate(rows):
        if idx <= opts.header_row_index:
            continue
        cells = list(row)
        if not any(_cell_str(c) for c in cells):
            continue

        def cell(key: str):
            i = m.get(key)
            return cells[i] if i is not None and i < len(cells) else None

        str_cells = [_cell_str(c) for c in cells]
        if opts.skip_value and opts.skip_value.casefold() in (s.casefold() for s in str_cells):
            result.skipped.append({"row_index": idx, "reason": "skip_value", "cells": str_cells})
            continue

        try:
            booked = parse_date_cell(cell("date"), opts.date_format)
        except (ValueError, TypeError):
            raw_date = _cell_str(cell("date"))
            reason = "reserverad" if "reserv" in raw_date.casefold() else "ogiltigt datum"
            result.skipped.append({"row_index": idx, "reason": reason, "cells": str_cells})
            continue

        try:
            if m.get("amount") is not None:
                amount = parse_amount(cell("amount"), opts.decimal_separator, opts.thousands_separator)
            else:
                inflow = cell("amount_in")
                outflow = cell("amount_out")
                has_in = inflow is not None and _cell_str(inflow)
                has_out = outflow is not None and _cell_str(outflow)
                if not has_in and not has_out:
                    # info-rader (t.ex. "Ingående saldo") ska inte bli 0-kronorsposter
                    result.skipped.append(
                        {"row_index": idx, "reason": "tomt belopp", "cells": str_cells}
                    )
                    continue
                amount = 0
                if has_in:
                    amount += abs(parse_amount(inflow, opts.decimal_separator, opts.thousands_separator))
                if has_out:
                    amount -= abs(parse_amount(outflow, opts.decimal_separator, opts.thousands_separator))
        except ValueError:
            result.skipped.append({"row_index": idx, "reason": "ogiltigt belopp", "cells": str_cells})
            continue

        if opts.invert_sign:
            amount = -amount

        balance = None
        if m.get("balance") is not None:
            try:
                balance = parse_amount(cell("balance"), opts.decimal_separator, opts.thousands_separator)
            except (ValueError, TypeError):
                balance = None

        desc = _cell_str(cell("description")) or "(okänd)"
        member = _cell_str(cell("member")) or None
        result.rows.append(
            ParsedRow(
                booked_date=booked,
                amount_ore=amount,
                description_raw=desc,
                description_norm=normalize_description(desc),
                balance_ore=balance,
                row_index=idx,
                member=member,
            )
        )
    return result
